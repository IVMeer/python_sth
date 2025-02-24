import openpyxl
from docx import Document

# 读取 Excel 文件
def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 读取第一个工作表
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

# 写入 Word 文件
def write_to_word(data, output_path):
    doc = Document()
    table = doc.add_table(rows=len(data), cols=len(data[0]))

    for i, row in enumerate(data):
        for j, cell_value in enumerate(row):
            table.cell(i, j).text = str(cell_value) if cell_value is not None else ''

    doc.save(output_path)

# 主函数
def excel_to_word(excel_path, word_path):
    data = read_excel(excel_path)
    write_to_word(data, word_path)
    print(f"数据已从 {excel_path} 转换为 {word_path}")

# 示例：将 Excel 文件转换为 Word 文件
excel_to_word('C:/Users/win11/Desktop/123.xlsx', 'C:/Users/win11/Desktop/t.docx')
